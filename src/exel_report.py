from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

from datetime import datetime
from pathlib import Path

def exel_report(audit):
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Audit Report'
    
    headers = ['Issue', 'Severity', 'Impact', 'Action']
    ws.append(headers)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_align = Alignment(horizontal='center', vertical='center')
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = header_align
    
    colors = {
        'CRITICAL': 'FF0000',
        'HIGH': 'FFC000',
        'MEDIUM': 'FFFF00',
        'LOW': '92D050',
        'DEFAULT': 'FFFFFF'
    }
    
    for item in audit:
        color = colors.get(item['severity'], colors['DEFAULT'])
        ws.append([item['issue'], item['severity'], item['impact'], item['action']])
        last_row = ws.max_row
        severity_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws.cell(row=last_row, column=2).fill = severity_fill
    
    ws.append([])
    ws.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws.append(['System:', 'AI Invoice Auditor v1.0'])

    BASE_DIR = Path(__file__).resolve().parent.parent
    REPORTS_DIR = BASE_DIR / 'reports'
    REPORTS_DIR.mkdir(exist_ok=True)
    
    file_path = REPORTS_DIR / 'audit_report.xlsx'
    wb.save(file_path)
    
    print(f'Exel report saved to: {file_path}')